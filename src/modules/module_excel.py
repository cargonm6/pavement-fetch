import os
import sys
import time

import csv

import openpyxl
import pandas
import win32com.client as wincl
from openpyxl.utils.dataframe import dataframe_to_rows

from src.modules.module_common import etr, to_hhmmss


def main(project_root):
    path_xls = os.path.abspath(project_root + '/res/Calculadora_PCI.xlsm')  # enter your file path
    path_csv = os.path.abspath(project_root + '/res/csv/ready/pci.csv')  # enter your file path
    sheet_name = "Data from LTPP"
    macro_name = "!Módulo2.PCI_sections"
    min_rows = 3

    if not os.path.isfile(path_xls):
        print(" (!) File \"%s\" does not exist" % path_xls)
        return
    if not os.path.isfile(path_csv):
        print(" (!) File \"%s\" does not exist" % path_xls)
        return

    # (1) Clear "Data from LTPP" tab
    print(" (1) Clearing data from Excel file \"%s\"" % path_xls)
    book = openpyxl.load_workbook(filename=path_xls, read_only=False, keep_vba=True)
    sheet = book[sheet_name]
    sheet.delete_rows(min_rows + 1, sheet.max_row - min_rows)

    # (2) Append CSV rows
    print(" (2) Appending rows from CSV \"%s\"" % path_csv)
    df = pandas.read_csv(path_csv, sep=';', decimal=',')
    df = df.apply(pandas.to_numeric, errors='ignore')
    df["OTHER"] = ""  # Solution for reserved characters in OpenPyXL
    length = len(df.index)
    count = 0
    time_for = 0
    for r in dataframe_to_rows(df, index=False, header=False):
        time_iter = time.time()
        count += 1
        sheet.append(r)
        time_est, time_for = etr(length, count, time_for, time_iter)
        if count < length:
            sys.stdout.write(
                "\r    - PCI %d/%d (ETR: %s)" % (count, length, time_est))
        else:
            sys.stdout.write(
                "\r    + PCI %d/%d (time: %s)\n" % (count, length, to_hhmmss(time_for)))
    book.save(path_xls)
    book.close()

    # (3) Execute PCI_sections macro
    print(" (3) Executing PCI calc (Excel macro \"%s\")..." % (os.path.basename(path_xls) + macro_name))
    excel_macro = wincl.DispatchEx("Excel.Application")
    workbook = excel_macro.Workbooks.Open(Filename=path_xls, ReadOnly=1)
    excel_macro.Application.Run(os.path.basename(path_xls) + macro_name)
    workbook.Save()
    excel_macro.Application.Quit()
    del workbook
    del excel_macro

    # (4) Copy results to CSV
    print(" (4) Copying results to CSV \"%s\")..." % path_csv)
    book = openpyxl.load_workbook(filename=path_xls, read_only=False, keep_vba=True)
    sheet = book[sheet_name]
    sheet.delete_rows(1, 2)
    with open(path_csv, 'w', newline='') as f:
        c = csv.writer(f, delimiter=';')
        for r in sheet.rows:
            c.writerow([cell.value for cell in r])
